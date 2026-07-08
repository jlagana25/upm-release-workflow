"""
split_se_ingest_forms.py — Split the SoundExchange metadata sheets into ISRC
Ingest Form workbooks (≤ MAX_ROWS_PER_WORKBOOK data rows each).

Consolidates the former per-entity scripts (8-Split_SE_Ingest_Form-MGB_NA_LLC.py
and ...-Z_TUNES_LLC.py — byte-identical except for one hardcoded path) into one
config-aware tool.

SoundExchange flow for the chosen month:
  • Inputs live in 2-STAGING/SoundExchange/ — the raw Domo metadata exports
    (Metadata/SoundExchange Universal Music - *.xlsx) and the ISRC Ingest Form
    template (ISRC Ingest Form.xlsx).
  • The generated "<base name> - Part N.xlsx" workbooks are written to the
    FINAL PACKAGING SoundExchange release folder
    (3-FINAL PACKAGING/Universal Production Music {Month} Release - SoundExchange).

    python3 split_se_ingest_forms.py --previous-month
    python3 split_se_ingest_forms.py --year 2026 --month 5 --part 1
    python3 split_se_ingest_forms.py --previous-month --only mgb

The split logic (template Form sheet, data starting at row 11, per-cell value +
style copy, template-row formatting) is unchanged from the originals.
"""

from __future__ import annotations

import argparse
import logging
from copy import copy
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from config import BASELINE_SPECIALS, context_from_cli_args

# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------
#
# SoundExchange flow:
#   • Inputs (raw Domo metadata exports + the ISRC Ingest Form template) live in
#     2-STAGING/SoundExchange/.
#   • The generated ingest-form workbooks ship from the FINAL PACKAGING
#     SoundExchange release folder.
# Both locations come from the release context (ctx.soundexchange_staging_dir /
# ctx.soundexchange_final_dir), so nothing here is hardcoded to a month.

TEMPLATE_NAME = "ISRC Ingest Form.xlsx"
FORM_SHEET = "Form"
START_ROW = 11
MAX_ROWS_PER_WORKBOOK = 9990

# entity → (ctx.partner_metadata key, output base name, --only token)
ENTITIES = [
    ("soundexchange_mgb",    "ISRC Ingest Form - MGB NA LLC", "mgb"),
    ("soundexchange_ztunes", "ISRC Ingest Form - Z TUNES LLC", "ztunes"),
]


def _resolve_template(ctx) -> Path:
    """Locate the ISRC Ingest Form template for this release.

    Preference order:
      1. This month's 2-STAGING/SoundExchange folder (where the template is
         placed alongside the raw exports).
      2. The baseline copy (2-STAGING/SoundExchange in the Specials baseline) —
         covers a month whose staging folder doesn't have it yet.
    Returns the month-folder path if neither exists, so the error message points
    at the expected location.
    """
    month_tpl = ctx.soundexchange_staging_dir / TEMPLATE_NAME
    baseline_tpl = (
        BASELINE_SPECIALS / "2-STAGING" / "SoundExchange" / TEMPLATE_NAME
    )
    for cand in (month_tpl, baseline_tpl):
        if cand.exists():
            return cand
    return month_tpl


# ---------------------------------------------------------------------------
# Cell / row helpers (unchanged from the originals)
# ---------------------------------------------------------------------------

def copy_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value

    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format

    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)

    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def clear_form_data(ws) -> None:
    for row in ws.iter_rows(min_row=START_ROW, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def apply_template_row_format(ws, row_num, max_col) -> None:
    template_row = START_ROW
    for col in range(1, max_col + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=row_num, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)
            dst.number_format = src.number_format


# ---------------------------------------------------------------------------
# Split one entity's sheet into ingest-form workbooks
# ---------------------------------------------------------------------------

def split_one(
    export_path: Path,
    base_name:   str,
    template_path: Path,
    output_dir:  Path,
    *,
    dry_run: bool = False,
    logger:  "Optional[logging.Logger]" = None,
) -> int:
    """Split a single SoundExchange export into ingest-form workbooks.

    Returns the number of part files written (or, in dry-run, would write).
    """
    emit = logger.info if logger else print

    export_wb = load_workbook(export_path, data_only=False)
    export_ws = export_wb.active

    rows = list(export_ws.iter_rows(min_row=2))   # skip export header
    max_col = export_ws.max_column

    if not dry_run:
        output_dir.mkdir(parents=True, exist_ok=True)

    part = 1
    for start in range(0, len(rows), MAX_ROWS_PER_WORKBOOK):
        chunk = rows[start:start + MAX_ROWS_PER_WORKBOOK]
        output_path = output_dir / f"{base_name} - Part {part}.xlsx"

        if dry_run:
            emit(f"    [WOULD WRITE] {output_path.name} — {len(chunk)} data row(s)")
            part += 1
            continue

        template_wb = load_workbook(template_path)
        form_ws = template_wb[FORM_SHEET]
        clear_form_data(form_ws)

        for row_offset, export_row in enumerate(chunk):
            target_row_num = START_ROW + row_offset
            apply_template_row_format(form_ws, target_row_num, max_col)
            for col_idx, source_cell in enumerate(export_row, start=1):
                target_cell = form_ws.cell(row=target_row_num, column=col_idx)
                copy_cell(source_cell, target_cell)

        template_wb.save(output_path)
        emit(f"    Saved {output_path.name} — {len(chunk)} data row(s)")
        part += 1

    if not rows:
        emit(f"    ⚠ {export_path.name} has no data rows — nothing written.")
    return part - 1


def run_soundexchange_split(
    ctx,
    *,
    dry_run: bool = False,
    logger:  "Optional[logging.Logger]" = None,
    only:    "Optional[str]" = None,
) -> bool:
    """
    Generate the SoundExchange ISRC Ingest Form workbooks for this release.

    Orchestrator-friendly entry point (driven as a Step 10 / final-packaging
    sub-phase): takes a ReleaseContext directly, logs via ``logger``, and
    honours ``dry_run``.  Reads the SoundExchange metadata sheets exported by
    Step 1 plus the ISRC Ingest Form template, and writes
    ``<base name> - Part N.xlsx`` workbooks into the FINAL PACKAGING
    SoundExchange folder (``ctx.soundexchange_final_dir``).

    Returns True on success.  Returns False if the template or any requested
    source sheet is missing — a real gap (run the Step 1 Domo export first).
    In ``dry_run`` a missing input is a warning, not a failure, so a preview
    never blocks.
    """
    emit  = logger.info    if logger else print
    warn  = logger.warning if logger else print
    error = logger.error   if logger else print

    only = (only or "").strip().lower()
    template   = _resolve_template(ctx)
    output_dir = ctx.soundexchange_final_dir
    emit(f"  Template: {template}")
    emit(f"  Output:   {output_dir}")

    if not template.exists():
        msg = (
            f"SoundExchange ISRC Ingest Form template not found: {template}  "
            f"(expected in 2-STAGING/SoundExchange, from the Specials baseline)."
        )
        if dry_run:
            warn("  ⚠ " + msg)
            emit("  [DRY RUN] Skipping SoundExchange preview (template absent).")
            return True
        error("  ✗ " + msg)
        return False

    total_files    = 0
    missing_sheets: list[str] = []
    for key, base_name, token in ENTITIES:
        if only and only != token:
            continue
        sheet = ctx.partner_metadata[key]
        emit(f"\n  {base_name}")
        emit(f"    source: {sheet}")
        if not sheet.exists():
            warn(
                "    ⚠ Source sheet not found — run the Step 1 Domo export "
                "for SoundExchange first."
            )
            missing_sheets.append(base_name)
            continue
        total_files += split_one(
            sheet, base_name, template, output_dir,
            dry_run=dry_run, logger=logger,
        )

    verb = "would write" if dry_run else "wrote"
    emit(f"\n  SoundExchange: {verb} {total_files} part file(s) → {output_dir}")

    if missing_sheets and not dry_run:
        error(
            "  ✗ SoundExchange incomplete — missing source sheet(s): "
            + ", ".join(missing_sheets)
        )
        return False
    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description=(
            "Split the SoundExchange final-packaging metadata sheets into "
            "ISRC Ingest Form workbooks (≤ "
            f"{MAX_ROWS_PER_WORKBOOK} rows each).  In a full pipeline run this "
            "also runs automatically as part of Step 10 (final packaging)."
        )
    )
    p.add_argument("--year",  type=int)
    p.add_argument("--month", type=int)
    p.add_argument("--part",  type=int, choices=[1, 2])
    p.add_argument("--previous-month", action="store_true",
                   help="Use the previous month's release (full-month run).")
    p.add_argument("--only", default=None,
                   help="Process only one entity: 'mgb' or 'ztunes'.")
    p.add_argument("--dry-run", action="store_true",
                   help="Report what would be written without creating files.")
    args = p.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )
    logger = logging.getLogger("soundexchange")

    ctx = context_from_cli_args(args)
    ok = run_soundexchange_split(
        ctx, dry_run=args.dry_run, logger=logger, only=args.only,
    )
    return 0 if ok else 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
