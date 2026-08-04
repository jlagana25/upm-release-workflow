"""Offline regression tests for release-integrity and destructive safeguards."""

from __future__ import annotations

import csv
import logging
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import final_metadata_verification as final_verify
import final_packaging
import domo_exports
import covers
import audio_conversion
import cleanup
import prune
import remediation
import unisync_automation
import verification
from audio_conversion import _flatten_mirrored_media
from cover_downloads import download_image_atomic
from final_packaging import CopyResult
from filesystem_names import resolve_label_dir
from logging_utils import get_logger
from release_manifest import ReleaseManifest
from soundmouse import metadata_filename, remove_stale_soundmouse_metadata
from split_se_ingest_forms import run_soundexchange_split, split_one


LOGGER = logging.getLogger("release_safety_test")


class ReleaseSafetyTests(unittest.TestCase):
    def test_wav_with_covers_missing_source_fails_real_run(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ctx = SimpleNamespace(specials_dir=Path(tmp) / "specials")
            self.assertFalse(
                covers.build_wav_with_covers_from_wav(ctx, False, LOGGER)
            )
            self.assertTrue(
                covers.build_wav_with_covers_from_wav(ctx, True, LOGGER)
            )

    def test_nbc_conversion_fails_when_wav_tree_is_empty(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            tail = (
                root / "3-FINAL PACKAGING"
                / "Universal Production Music May 2026 Release - NBC"
                / "Music"
            )
            wav = tail / "WAV"
            wav.mkdir(parents=True)
            ctx = SimpleNamespace(
                month_display_folder="May 2026",
                partner_dirs={
                    "nbc_wav_music": wav,
                    "nbc_mp3_music": tail / "MP3",
                },
            )
            with patch.object(
                audio_conversion.shutil, "which",
                side_effect=lambda name: f"/usr/bin/{name}",
            ):
                self.assertFalse(
                    audio_conversion.convert_nbc_wav_to_mp3(
                        ctx, False, False, LOGGER
                    )
                )

    def test_grouped_packaging_walks_shared_source_once(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            source.mkdir()
            (source / "track.wav").write_bytes(b"audio")
            ops = [
                final_packaging.CopyOp("one", source, root / "one"),
                final_packaging.CopyOp("two", source, root / "two"),
            ]
            real_walk = final_packaging.os.walk
            calls = 0

            def counted_walk(path):
                nonlocal calls
                calls += 1
                return real_walk(path)

            with patch.object(final_packaging.os, "walk", new=counted_walk):
                results = final_packaging._run_grouped_ops(
                    ops, dry_run=False, overwrite=False, logger=LOGGER
                )
            self.assertEqual(calls, 1)
            self.assertTrue(all(result.ok for result in results))
            self.assertTrue((root / "one" / "track.wav").is_file())
            self.assertTrue((root / "two" / "track.wav").is_file())

    def test_tunesat_materializes_only_metadata_keepers_from_both_sources(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            music = root / "specials" / "1-ORIGINAL" / "Music"
            us = music / "MP3" / "MEDIA" / "US" / "ALB - Title"
            exus = (
                music / "Ex-US (MP3)" / "MEDIA"
                / "BTV" / "EX - Title"
            )
            us.mkdir(parents=True)
            exus.mkdir(parents=True)
            (us / "US_001.mp3").write_bytes(b"us")
            (us / "EXTRA_999.mp3").write_bytes(b"extra")
            (exus / "EX_002.mp3").write_bytes(b"ex")
            metadata = root / "tunesat.csv"
            metadata.write_text(
                "File Name\nUS_001.mp3\nEX_002.mp3\n", encoding="utf-8"
            )
            target = root / "tunesat"
            target.mkdir()
            ctx = SimpleNamespace(
                specials_dir=root / "specials",
                cleanup_metadata_csv=metadata,
                cleanup_target_folder=target,
            )
            self.assertTrue(
                cleanup.remove_non_maintracks(
                    ctx, False, True, LOGGER
                )
            )
            delivered = {path.name for path in target.rglob("*.mp3")}
            self.assertEqual(delivered, {"US_001.mp3", "EX_002.mp3"})

    def test_release_manifest_reloads_changed_export(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "tracklist.csv"
            manifest = ReleaseManifest()
            path.write_text("name\nold\n", encoding="utf-8")
            self.assertEqual(manifest.read(path).iloc[0, 0], "old")
            path.write_text("name\nnew-value\n", encoding="utf-8")
            self.assertEqual(manifest.read(path).iloc[0, 0], "new-value")

    def test_remediation_uses_current_in_memory_findings(self) -> None:
        finding = {
            "Type": "COVERS",
            "Filename": "cover.jpg",
            "Source CSV": "US Tracklist",
            "Reason": "missing",
        }
        calls = 0

        def verify(_ctx, _dry_run, _logger, *, findings_out=None):
            nonlocal calls
            calls += 1
            if calls == 1:
                findings_out.extend([finding])
                return False
            return True

        with (
            patch.object(verification, "verify_all_files", new=verify),
            patch.object(remediation, "remediate_from_rows") as remediate,
        ):
            self.assertEqual(
                remediation.verify_and_remediate_loop(
                    SimpleNamespace(), 2, False, False, False, LOGGER,
                    run_domo=False,
                ),
                (True, 0),
            )
        self.assertEqual(remediate.call_args.args[1], [finding])

    def test_atomic_cover_download_validates_before_replace(self) -> None:
        class Response:
            def __init__(self, payload: bytes):
                self.payload = payload

            def raise_for_status(self):
                return None

            def iter_content(self, _size):
                yield self.payload

        class Session:
            def __init__(self, payload: bytes):
                self.payload = payload

            def get(self, *_args, **_kwargs):
                return Response(self.payload)

        with tempfile.TemporaryDirectory() as tmp:
            destination = Path(tmp) / "cover.png"
            destination.write_bytes(b"old")
            with self.assertRaises(ValueError):
                download_image_atomic(
                    "https://example.invalid/cover",
                    destination,
                    timeout=1,
                    session=Session(b"not-image"),
                )
            self.assertEqual(destination.read_bytes(), b"old")
            png = b"\x89PNG\r\n\x1a\n" + b"valid-enough"
            download_image_atomic(
                "https://example.invalid/cover",
                destination,
                timeout=1,
                session=Session(png),
            )
            self.assertEqual(destination.read_bytes(), png)

    def test_whitespace_label_is_resolved_and_kept_by_prune(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            track = root / "BTV " / "ALB - Title" / "track.mp3"
            track.parent.mkdir(parents=True)
            track.write_bytes(b"audio")

            self.assertEqual(resolve_label_dir(root, "BTV"), root / "BTV ")
            extras, _junk, albums = prune._scan_tree(
                root, {("btv", "ALB", "track.mp3")}, set(), False
            )
            self.assertEqual(extras, [])
            self.assertTrue(albums)

    def test_flatten_requires_overwrite_and_preserves_existing_tree(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            wav = Path(tmp) / "WAV"
            old = wav / "MEDIA" / "old.wav"
            nested = wav / "_Specials" / "path" / "MEDIA" / "new.wav"
            old.parent.mkdir(parents=True)
            nested.parent.mkdir(parents=True)
            old.write_bytes(b"old")
            nested.write_bytes(b"new")

            self.assertFalse(
                _flatten_mirrored_media(wav, LOGGER, dry_run=False)
            )
            self.assertEqual(old.read_bytes(), b"old")
            self.assertTrue(nested.exists())

            self.assertTrue(
                _flatten_mirrored_media(
                    wav, LOGGER, dry_run=False, overwrite=True
                )
            )
            self.assertFalse(old.exists())
            self.assertEqual((wav / "MEDIA" / "new.wav").read_bytes(), b"new")

    def test_flatten_restores_previous_tree_when_install_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            wav = Path(tmp) / "WAV"
            target = wav / "MEDIA"
            old = target / "old.wav"
            nested = wav / "_Specials" / "path" / "MEDIA"
            new = nested / "new.wav"
            old.parent.mkdir(parents=True)
            new.parent.mkdir(parents=True)
            old.write_bytes(b"old")
            new.write_bytes(b"new")
            original_rename = Path.rename

            def fail_nested(path: Path, destination: Path):
                if path == nested:
                    raise OSError("simulated move failure")
                return original_rename(path, destination)

            with patch.object(Path, "rename", new=fail_nested):
                self.assertFalse(
                    _flatten_mirrored_media(
                        wav, LOGGER, dry_run=False, overwrite=True
                    )
                )
            self.assertEqual(old.read_bytes(), b"old")
            self.assertEqual(new.read_bytes(), b"new")

    def test_flatten_ignores_media_directories_inside_flat_target(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            wav = Path(tmp) / "WAV"
            track = wav / "MEDIA" / "label" / "MEDIA" / "track.wav"
            track.parent.mkdir(parents=True)
            track.write_bytes(b"audio")

            self.assertTrue(
                _flatten_mirrored_media(wav, LOGGER, dry_run=False)
            )
            self.assertEqual(track.read_bytes(), b"audio")

    def test_missing_packaging_source_is_failure(self) -> None:
        result = CopyResult("test", Path("missing"), Path("destination"))
        result.source_missing = True
        self.assertFalse(result.ok)

    def test_unisync_zero_progress_is_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "tracks.csv"
            source.write_text("workAudioId\n1\n", encoding="utf-8")
            job = {
                "name": "US WAV",
                "cache_path": str(root / "cache"),
                "client_path": str(root / "client"),
                "csv": str(source),
            }
            with (
                patch.object(
                    unisync_automation,
                    "_expected_output_filenames",
                    return_value={"track.wav"},
                ),
                patch.object(
                    unisync_automation, "_present_filenames", return_value=set()
                ),
                patch.object(
                    unisync_automation,
                    "_drive_unisync_for_csv",
                    return_value=unisync_automation.STATUS_OK,
                ),
            ):
                self.assertEqual(
                    unisync_automation._run_single_job(job, False, LOGGER),
                    unisync_automation.STATUS_FAILED,
                )

    def test_final_verification_fails_for_missing_required_media(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            check = final_verify.Check(
                "Required partner", root / "source.csv", root / "missing-media"
            )
            ctx = SimpleNamespace(
                specials_dir=root, month_display_folder="May 2026"
            )
            with (
                patch.object(final_verify, "_build_checks", return_value=[check]),
                patch.object(final_verify, "EXPORTS_DIR", root / "reports"),
            ):
                self.assertFalse(
                    final_verify.verify_final_packaging_metadata(
                        ctx, LOGGER, dry_run=False
                    )
                )

    def test_final_verification_rejects_empty_expected_set(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.csv"
            source.write_text("Filename\n", encoding="utf-8")
            media = root / "media"
            media.mkdir()
            check = final_verify.Check("Required", source, media)
            ctx = SimpleNamespace(
                specials_dir=root, month_display_folder="May 2026"
            )
            with (
                patch.object(final_verify, "_build_checks", return_value=[check]),
                patch.object(final_verify, "EXPORTS_DIR", root / "reports"),
            ):
                self.assertFalse(
                    final_verify.verify_final_packaging_metadata(
                        ctx, LOGGER, dry_run=False
                    )
                )

    def test_verification_dry_run_does_not_write_report(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            specials = root / "specials"
            specials.mkdir()
            report = root / "reports" / "missing.csv"
            ctx = SimpleNamespace(
                specials_dir=specials,
                missing_report_csv=report,
            )
            missing = [{field: "x" for field in verification.REPORT_FIELDS}]
            with (
                patch.object(verification, "_verify_us", return_value=missing),
                patch.object(verification, "_verify_exus", return_value=[]),
                patch.object(verification, "_verify_japan", return_value=[]),
            ):
                self.assertFalse(verification.verify_all_files(ctx, True, LOGGER))
            self.assertFalse(report.exists())

    def test_prune_dry_run_can_suppress_report_write(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            report = root / "reports" / "missing.csv"
            ctx = SimpleNamespace(
                specials_dir=root / "specials",
                missing_report_csv=report,
                month_display="May 2026",
            )
            with patch.object(prune, "_tree_specs", return_value=[]):
                self.assertEqual(
                    prune.prune_music_trees(
                        ctx, prune.PRUNE_REPORT, LOGGER, write_report=False
                    ),
                    (0, 0),
                )
            self.assertFalse(report.parent.exists())

    def test_console_only_logger_writes_no_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            log_dir = Path(tmp) / "logs"
            _logger, log_path = get_logger(
                2099, 1, 1, log_dir=log_dir, write_file=False
            )
            self.assertIsNone(log_path)
            self.assertFalse(log_dir.exists())

    def test_domo_csv_conversion_replaces_only_after_success(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "download.xlsx"
            destination = root / "export.csv"
            destination.write_text("old\nvalue\n", encoding="utf-8")

            workbook = Workbook()
            workbook.active.append(["Header"])
            workbook.active.append(["new"])
            workbook.save(source)
            workbook.close()

            with patch("pandas.DataFrame.to_csv", side_effect=OSError("disk full")):
                with self.assertRaises(OSError):
                    domo_exports._xlsx_to_csv(source, destination, LOGGER)
            self.assertEqual(
                destination.read_text(encoding="utf-8"), "old\nvalue\n"
            )
            self.assertTrue(source.exists())

            domo_exports._xlsx_to_csv(source, destination, LOGGER)
            self.assertIn("new", destination.read_text(encoding="utf-8-sig"))
            self.assertFalse(source.exists())

    def test_domo_xlsx_export_preserves_old_until_validated(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            destination = root / "export.xlsx"

            old = Workbook()
            old.active["A1"] = "old"
            old.save(destination)
            old.close()

            invalid = root / "invalid.xlsx"
            invalid.write_bytes(b"not a workbook")
            page = SimpleNamespace(url="https://domo.test/kpis/details/123")
            ctx = SimpleNamespace(previous_month=True)
            card = {
                "key": "test",
                "card_id": "123",
                "description": "test card",
                "format": "xlsx",
            }
            with (
                patch.object(domo_exports, "_navigate_to_card"),
                patch.object(domo_exports, "_apply_previous_month_preset"),
                patch.object(
                    domo_exports,
                    "_trigger_excel_download",
                    return_value=invalid,
                ),
            ):
                with self.assertRaises(RuntimeError):
                    domo_exports._export_card(
                        page, card, destination, ctx, LOGGER
                    )
            retained = load_workbook(destination, read_only=True)
            self.assertEqual(retained.active["A1"].value, "old")
            retained.close()

            valid = root / "valid.xlsx"
            new = Workbook()
            new.active["A1"] = "new"
            new.save(valid)
            new.close()
            with (
                patch.object(domo_exports, "_navigate_to_card"),
                patch.object(domo_exports, "_apply_previous_month_preset"),
                patch.object(
                    domo_exports,
                    "_trigger_excel_download",
                    return_value=valid,
                ),
            ):
                domo_exports._export_card(
                    page, card, destination, ctx, LOGGER
                )
            replaced = load_workbook(destination, read_only=True)
            self.assertEqual(replaced.active["A1"].value, "new")
            replaced.close()

    def test_domo_followup_cards_share_one_browser_session(self) -> None:
        launches = 0

        class Browser:
            def new_context(self, **_kwargs):
                return self

            def new_page(self):
                return object()

            def close(self):
                return None

        class Chromium:
            def launch(self, **_kwargs):
                nonlocal launches
                launches += 1
                return Browser()

        class Playwright:
            chromium = Chromium()

        class Manager:
            def __enter__(self):
                return Playwright()

            def __exit__(self, *_args):
                return False

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            ctx = SimpleNamespace(
                release_start="2026-05-01",
                release_end="2026-05-14",
                tracklist_token="test",
                previous_month=True,
                us_tracklist_csv=root / "us.csv",
            )
            seed = {
                "key": "seed",
                "card_id": "1",
                "description": "Seed",
                "output_fn": lambda _ctx: root / "seed.csv",
            }
            followup = {
                "key": "followup",
                "card_id": "2",
                "description": "Follow-up",
                "output_fn": lambda _ctx: root / "followup.csv",
            }

            def export(_page, _card, output, _ctx, _logger):
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_text("ok", encoding="utf-8")

            with (
                patch.object(domo_exports, "sync_playwright", return_value=Manager()),
                patch.object(domo_exports, "_authenticate"),
                patch.object(domo_exports, "_export_card", new=export),
                patch.object(domo_exports, "TEMP_DOWNLOAD_DIR", root / "downloads"),
            ):
                results = domo_exports.run_domo_exports(
                    ctx,
                    False,
                    LOGGER,
                    only_keys=["us_tracklist"],
                    extra_cards=[seed],
                    followup_cards=lambda: [followup],
                )
            self.assertEqual(launches, 1)
            self.assertEqual(
                results,
                {"us_tracklist": "ok", "seed": "ok", "followup": "ok"},
            )

    def test_soundexchange_removes_stale_parts_after_success(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            export = root / "export.xlsx"
            template = root / "template.xlsx"
            output = root / "output"

            workbook = Workbook()
            workbook.active.append(["Header"])
            workbook.active.append(["value"])
            workbook.save(export)
            workbook.close()

            workbook = Workbook()
            workbook.active.title = "Form"
            workbook.active.cell(row=11, column=1, value="template")
            workbook.save(template)
            workbook.close()

            output.mkdir()
            stale = output / "Entity - Part 2.xlsx"
            lookalike = output / "Entity - Part notes.xlsx"
            stale.write_bytes(b"stale")
            lookalike.write_bytes(b"keep")
            self.assertEqual(split_one(export, "Entity", template, output), 1)
            self.assertTrue((output / "Entity - Part 1.xlsx").is_file())
            self.assertFalse(stale.exists())
            self.assertTrue(lookalike.exists())
            result = load_workbook(output / "Entity - Part 1.xlsx")
            self.assertEqual(result["Form"].cell(row=11, column=1).value, "value")
            result.close()

    def test_soundexchange_rejects_unknown_only_value(self) -> None:
        self.assertFalse(
            run_soundexchange_split(SimpleNamespace(), only="unknown")
        )

    def test_soundmouse_removes_only_unselected_generated_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            metadata = Path(tmp)
            selected = metadata / metadata_filename("01")
            stale = metadata / metadata_filename("02")
            unrelated = metadata / "notes.xlsx"
            lookalike = metadata / "SoundMouseMetadata notes.xlsx"
            for path in (selected, stale, unrelated, lookalike):
                path.write_bytes(b"x")
            self.assertEqual(
                remove_stale_soundmouse_metadata(
                    metadata, ["01"], False, LOGGER
                ),
                1,
            )
            self.assertTrue(selected.exists())
            self.assertFalse(stale.exists())
            self.assertTrue(unrelated.exists())
            self.assertTrue(lookalike.exists())


if __name__ == "__main__":
    unittest.main()
