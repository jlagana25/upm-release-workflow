"""
soundmouse.py — Step 16: Build the SoundMouse release delivery.

The step exports the SoundMouse tracklist and bucket from Domo, creates one
workflow-period directory with Covers/Metadata/MEDIA children, downloads WAVs
with UniSync and cover art from the tracklist, then exports only the metadata
workbooks named by the bucket card.  Raw Domo ActivationRange values never
control delivery-directory naming.

All browser and GUI dependencies remain lazy so this module imports headless.
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import shutil
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from config import (
    SOUNDMOUSE_BASE,
    SOUNDMOUSE_DOMO_CARDS,
    SOUNDMOUSE_DOMO_PAGE_ID,
    DOMO_PROFILE_DIR,
    UPM_CACHE_WAV,
    ReleaseContext,
    context_from_cli_args,
)
from auth_manager import private_creation_umask, secure_private_directory
from tracklist_columns import (
    POSSIBLE_COVER_COLS,
    POSSIBLE_FILENAME_COLS,
    POSSIBLE_URL_COLS,
    _find_column,
)


ACTIVATION_RANGE_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2}_to_\d{4}-\d{2}-\d{2}$"
)
# code -> (territory label, territory set).  The set lets us recognize either
# a bucket label ("02 UK DE SE OZ") or a raw Territory List value.
METADATA_SHEETS: dict[str, tuple[str, frozenset[str]]] = {
    "01": ("ALL",         frozenset({"US", "UK", "DE", "SE", "OZ"})),
    "02": ("UK DE SE OZ", frozenset({"UK", "DE", "SE", "OZ"})),
    "03": ("US DE SE OZ", frozenset({"US", "DE", "SE", "OZ"})),
    "04": ("DE SE OZ",    frozenset({"DE", "SE", "OZ"})),
    "05": ("UK SE OZ",    frozenset({"UK", "SE", "OZ"})),
    "06": ("DE SE",       frozenset({"DE", "SE"})),
    "07": ("US OZ",       frozenset({"US", "OZ"})),
    "08": ("OZ",          frozenset({"OZ"})),
    "09": ("SE",          frozenset({"SE"})),
    "10": ("US",          frozenset({"US"})),
}


def metadata_filename(code: str) -> str:
    label = METADATA_SHEETS[code][0]
    return f"SoundMouseMetadata {code} - {label}.xlsx"


def metadata_csv_filename(code: str) -> str:
    return Path(metadata_filename(code)).with_suffix(".csv").name


def convert_soundmouse_csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """Convert a Domo CSV export into a clean, upload-compatible XLSX.

    Every CSV field is written as literal text, so leading zeroes, long IDs,
    and values beginning with ``=`` survive unchanged. The XLSX package is then
    normalized to use Excel's shared-string table, which the SoundMouse uploader
    requires, without automating Excel or touching any open workbooks.
    """
    if not csv_path.is_file():
        raise FileNotFoundError(f"SoundMouse CSV export is missing: {csv_path}")

    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if len(rows) < 2 or not rows[0]:
        raise ValueError(f"SoundMouse CSV has no data rows: {csv_path}")
    expected_columns = len(rows[0])
    for row_number, row in enumerate(rows[1:], start=2):
        if len(row) != expected_columns:
            raise ValueError(
                f"SoundMouse CSV row {row_number} has {len(row)} columns; "
                f"expected {expected_columns}: {csv_path}"
            )

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = xlsx_path.with_name(f".{xlsx_path.name}.from-csv.tmp.xlsx")
    generated = temporary.with_name(f".{xlsx_path.name}.generated.tmp.xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell

        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet("Metadata")
        for row in rows:
            cells = []
            for value in row:
                cell = WriteOnlyCell(worksheet, value=value)
                cell.data_type = "s"
                cells.append(cell)
            worksheet.append(cells)
        workbook.save(generated)
        workbook.close()

        _rewrite_inline_strings_as_shared(generated, temporary)
        _assert_soundmouse_xlsx_compatibility(temporary)
        temporary.replace(xlsx_path)
        _assert_soundmouse_xlsx_compatibility(xlsx_path)
    finally:
        if generated.exists():
            generated.unlink()
        if temporary.exists():
            temporary.unlink()


def _rewrite_inline_strings_as_shared(source: Path, destination: Path) -> None:
    """Rewrite openpyxl inline strings into an OOXML shared-string table."""
    spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationships_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_types_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("", spreadsheet_ns)
    ET.register_namespace("", relationships_ns)

    shared_values: list[str] = []
    shared_indexes: dict[str, int] = {}
    total_strings = 0

    with ZipFile(source) as input_package:
        members = {name: input_package.read(name) for name in input_package.namelist()}

    for name, payload in list(members.items()):
        if not (name.startswith("xl/worksheets/sheet") and name.endswith(".xml")):
            continue
        root = ET.fromstring(payload)
        changed = False
        for cell in root.iter(f"{{{spreadsheet_ns}}}c"):
            if cell.get("t") != "inlineStr":
                continue
            inline = cell.find(f"{{{spreadsheet_ns}}}is")
            value = "" if inline is None else "".join(inline.itertext())
            index = shared_indexes.get(value)
            if index is None:
                index = len(shared_values)
                shared_indexes[value] = index
                shared_values.append(value)
            total_strings += 1
            for child in list(cell):
                cell.remove(child)
            cell.set("t", "s")
            ET.SubElement(cell, f"{{{spreadsheet_ns}}}v").text = str(index)
            changed = True
        if changed:
            members[name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    if not total_strings:
        raise ValueError(f"No string cells found while converting {source.name}")

    shared_root = ET.Element(
        f"{{{spreadsheet_ns}}}sst",
        {"count": str(total_strings), "uniqueCount": str(len(shared_values))},
    )
    for value in shared_values:
        item = ET.SubElement(shared_root, f"{{{spreadsheet_ns}}}si")
        text = ET.SubElement(item, f"{{{spreadsheet_ns}}}t")
        if value[:1].isspace() or value[-1:].isspace():
            text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        text.text = value
    members["xl/sharedStrings.xml"] = ET.tostring(
        shared_root, encoding="utf-8", xml_declaration=True
    )

    relationships = ET.fromstring(members["xl/_rels/workbook.xml.rels"])
    relationship_type = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
    )
    if not any(item.get("Type") == relationship_type for item in relationships):
        used_ids = {item.get("Id", "") for item in relationships}
        next_id = 1
        while f"rId{next_id}" in used_ids:
            next_id += 1
        ET.SubElement(
            relationships,
            f"{{{relationships_ns}}}Relationship",
            {"Id": f"rId{next_id}", "Type": relationship_type, "Target": "sharedStrings.xml"},
        )
    members["xl/_rels/workbook.xml.rels"] = ET.tostring(
        relationships, encoding="utf-8", xml_declaration=True
    )

    content_types = ET.fromstring(members["[Content_Types].xml"])
    part_name = "/xl/sharedStrings.xml"
    if not any(item.get("PartName") == part_name for item in content_types):
        ET.SubElement(
            content_types,
            f"{{{content_types_ns}}}Override",
            {
                "PartName": part_name,
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sharedStrings+xml"
                ),
            },
        )
    members["[Content_Types].xml"] = ET.tostring(
        content_types, encoding="utf-8", xml_declaration=True
    )

    with ZipFile(destination, "w", compression=ZIP_DEFLATED) as output_package:
        for name, payload in members.items():
            output_package.writestr(name, payload)


def _assert_soundmouse_xlsx_compatibility(path: Path) -> None:
    """Reject packages using string storage known to fail SoundMouse upload."""
    with ZipFile(path) as package:
        names = set(package.namelist())
        if "xl/sharedStrings.xml" not in names:
            raise ValueError(f"{path.name} has no Excel shared-string table")
        for name in names:
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                if b't="inlineStr"' in package.read(name):
                    raise ValueError(
                        f"{path.name} still contains inline strings rejected by SoundMouse"
                    )


def _file_key(value: object) -> str:
    """Case-insensitive leaf filename key; metadata paths never control IO."""
    return Path(str(value or "").strip()).name.casefold()


def _disk_file_keys(root: Path) -> set[str]:
    if not root.is_dir():
        return set()
    return {
        path.name.casefold()
        for path in root.rglob("*")
        if path.is_file() and path.stat().st_size > 0
    }


def validate_soundmouse_delivery(
    metadata_paths: list[Path],
    media_root: Path,
    covers_root: Path,
    report_path: Path,
    dry_run: bool,
    logger: logging.Logger,
    additional_media_roots: tuple[Path, ...] = (),
    additional_cover_roots: tuple[Path, ...] = (),
) -> bool:
    """Confirm every audio and cover named by the period metadata is present.

    Expected names are the union across every bucket-selected workbook.
    Matching is recursive and case-insensitive by exact leaf filename.  A CSV
    report is always written on a real run (header-only when clean), making the
    final gate auditable and preventing a stale prior report from lingering.
    """
    if dry_run:
        logger.info(
            f"  [DRY RUN] Would validate metadata audio against {media_root}"
        )
        logger.info(
            f"  [DRY RUN] Would validate metadata covers against {covers_root}"
        )
        logger.info(f"  [DRY RUN] Validation report: {report_path}")
        return True

    from openpyxl import load_workbook

    expected_audio: dict[str, set[str]] = {}
    expected_covers: dict[str, set[str]] = {}
    expected_names: dict[str, str] = {}
    errors: list[dict[str, str]] = []

    def add_expected(
        destination: dict[str, set[str]], raw: object, source: Path
    ) -> None:
        key = _file_key(raw)
        if key:
            destination.setdefault(key, set()).add(source.name)
            expected_names.setdefault(
                key, Path(str(raw or "").strip()).name
            )

    for metadata_path in metadata_paths:
        if not metadata_path.is_file():
            errors.append({
                "Type": "METADATA",
                "Filename": metadata_path.name,
                "Metadata Workbooks": metadata_path.name,
                "Expected Root": str(metadata_path.parent),
                "Problem": "Metadata workbook is missing",
            })
            continue
        try:
            workbook = load_workbook(
                metadata_path, read_only=True, data_only=False
            )
            try:
                recognized_sheet = False
                for worksheet in workbook.worksheets:
                    rows = worksheet.iter_rows(values_only=True)
                    header = next(rows, None)
                    if not header:
                        continue
                    headers = [str(value or "").strip() for value in header]
                    audio_col = _find_column(headers, POSSIBLE_FILENAME_COLS)
                    cover_col = _find_column(headers, POSSIBLE_COVER_COLS)
                    if not audio_col or not cover_col:
                        continue
                    recognized_sheet = True
                    audio_index = headers.index(audio_col)
                    cover_index = headers.index(cover_col)
                    for row in rows:
                        if audio_index < len(row):
                            add_expected(
                                expected_audio, row[audio_index], metadata_path
                            )
                        if cover_index < len(row):
                            add_expected(
                                expected_covers, row[cover_index], metadata_path
                            )
                if not recognized_sheet:
                    errors.append({
                        "Type": "METADATA",
                        "Filename": metadata_path.name,
                        "Metadata Workbooks": metadata_path.name,
                        "Expected Root": str(metadata_path),
                        "Problem": (
                            "No worksheet contains both an audio filename "
                            "and cover-art filename column"
                        ),
                    })
            finally:
                workbook.close()
        except Exception as exc:
            errors.append({
                "Type": "METADATA",
                "Filename": metadata_path.name,
                "Metadata Workbooks": metadata_path.name,
                "Expected Root": str(metadata_path),
                "Problem": f"Could not read workbook: {exc}",
            })

    present_audio = set().union(*(
        _disk_file_keys(root) for root in (media_root, *additional_media_roots)
    ))
    present_covers = set().union(*(
        _disk_file_keys(root) for root in (covers_root, *additional_cover_roots)
    ))
    missing_audio = sorted(set(expected_audio) - present_audio)
    missing_covers = sorted(set(expected_covers) - present_covers)

    for filename in missing_audio:
        errors.append({
            "Type": "AUDIO",
            "Filename": expected_names[filename],
            "Metadata Workbooks": "; ".join(sorted(expected_audio[filename])),
            "Expected Root": str(media_root),
            "Problem": "Referenced audio file is missing",
        })
    for filename in missing_covers:
        errors.append({
            "Type": "COVER",
            "Filename": expected_names[filename],
            "Metadata Workbooks": "; ".join(sorted(expected_covers[filename])),
            "Expected Root": str(covers_root),
            "Problem": "Referenced cover file is missing",
        })

    report_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "Type", "Filename", "Metadata Workbooks", "Expected Root", "Problem"
    ]
    with report_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(errors)

    logger.info(
        f"  SoundMouse validation: {len(expected_audio)} audio, "
        f"{len(expected_covers)} covers referenced by metadata."
    )
    logger.info(
        f"  Present: {len(present_audio)} files under MEDIA, "
        f"{len(present_covers)} files under Covers."
    )
    logger.info(f"  Validation report: {report_path}")
    if errors:
        logger.error(
            f"  ✗ SoundMouse validation failed: {len(missing_audio)} missing "
            f"audio, {len(missing_covers)} missing covers, and "
            f"{len(errors) - len(missing_audio) - len(missing_covers)} "
            "metadata error(s)."
        )
        for item in errors[:25]:
            logger.error(
                f"     {item['Type']}: {item['Filename']} — {item['Problem']}"
            )
        if len(errors) > 25:
            logger.error(f"     … and {len(errors) - 25} more; see report.")
        return False

    logger.info("  ✓ SoundMouse validation passed — all metadata audio and covers exist.")
    return True


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError(f"CSV has no header row: {path}")
        rows = [dict(row) for row in reader]
        return list(reader.fieldnames), rows


def activation_ranges_from_tracklist(path: Path) -> list[str]:
    """Return safe, distinct ActivationRange values in first-seen order."""
    fields, rows = _read_csv(path)
    col = _find_column(fields, ["ACTIVATIONRANGE"])
    if not col:
        raise ValueError(f"SoundMouse tracklist has no ActivationRange column: {path}")

    ranges: list[str] = []
    for row in rows:
        value = str(row.get(col, "")).strip()
        if not value:
            continue
        if not ACTIVATION_RANGE_RE.fullmatch(value):
            raise ValueError(
                f"Unsafe/invalid ActivationRange {value!r} in {path}; expected "
                "YYYY-MM-DD_to_YYYY-MM-DD."
            )
        if value not in ranges:
            ranges.append(value)
    return ranges


def create_soundmouse_directories(
    release_directory: Path,
    dry_run: bool,
    logger: logging.Logger,
) -> Path:
    """Create the one delivery root resolved by the workflow period.

    Domo's ``ActivationRange`` values describe upstream activation windows and
    may not match this workflow's Part 1/Part 2/full-month boundaries.  They
    therefore never control the delivery directory name.
    """
    for child in (
        release_directory,
        release_directory / "Covers",
        release_directory / "Metadata",
        release_directory / "MEDIA",
    ):
        if dry_run:
            logger.info(f"  [DRY RUN] Would create: {child}")
        else:
            child.mkdir(parents=True, exist_ok=True)
    logger.info(
        f"  {'Would prepare' if dry_run else 'Prepared'}: {release_directory}"
    )
    return release_directory


def _territory_set(value: str) -> frozenset[str]:
    tokens = set(re.findall(r"[A-Z]+", value.upper()))
    return frozenset(tokens & {"US", "UK", "DE", "SE", "OZ"})


def metadata_codes_from_bucket(path: Path) -> list[str]:
    """Recognize bucket codes/names or exact territory combinations."""
    _fields, rows = _read_csv(path)
    found: set[str] = set()
    by_territory = {territories: code for code, (_label, territories) in METADATA_SHEETS.items()}

    for row in rows:
        for raw in row.values():
            value = str(raw or "").strip()
            if not value:
                continue
            code_match = re.search(r"^\s*(0[1-9]|10)(?:\D|$)", value)
            if code_match:
                found.add(code_match.group(1))
            territories = _territory_set(value)
            if territories in by_territory:
                found.add(by_territory[territories])
            if value.upper() == "ALL":
                found.add("01")
    return sorted(found, key=int)


def _domo_configs(
    ctx: ReleaseContext,
    codes: list[str] | None = None,
    *,
    metadata_dir: Path | None = None,
) -> list[dict]:
    if codes is None:
        return [
            {
                "key": "soundmouse_tracklist",
                "card_id": SOUNDMOUSE_DOMO_CARDS["tracklist"],
                "page_id": SOUNDMOUSE_DOMO_PAGE_ID,
                "description": "SoundMouse Tracklist",
                "output_fn": lambda _ctx: ctx.soundmouse_tracklist_csv,
            },
            {
                "key": "soundmouse_bucket",
                "card_id": SOUNDMOUSE_DOMO_CARDS["bucket"],
                "page_id": SOUNDMOUSE_DOMO_PAGE_ID,
                "description": "SoundMouse Bucket",
                "output_fn": lambda _ctx: ctx.soundmouse_bucket_csv,
            },
        ]

    output_dir = metadata_dir or (ctx.soundmouse_release_dir / "Metadata")
    return [
        {
            "key": f"soundmouse_metadata_{code}",
            "card_id": SOUNDMOUSE_DOMO_CARDS[code],
            "page_id": SOUNDMOUSE_DOMO_PAGE_ID,
            "description": f"SoundMouse Metadata {code}",
            "output_fn": (
                lambda _ctx, c=code: output_dir / metadata_csv_filename(c)
            ),
            "format": "csv",
            "download_format": "csv",
            "xlsx_output_fn": (
                lambda _ctx, c=code: output_dir / metadata_filename(c)
            ),
        }
        for code in codes
    ]


def _export_domo_cards(
    ctx: ReleaseContext,
    cards: list[dict],
    dry_run: bool,
    logger: logging.Logger,
) -> bool:
    if dry_run:
        for card in cards:
            logger.info(
                f"  [DRY RUN] Would export {card['description']}: "
                f"{card['output_fn'](ctx)}"
            )
        return True

    # domo_exports owns the known-good Domo interaction.  Its Playwright import
    # is deliberately initialized here, inside the real-run function.
    import domo_exports as domo

    domo.TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    domo._require_playwright()
    ok = True
    secure_private_directory(DOMO_PROFILE_DIR, recursive=True)
    with domo.sync_playwright() as playwright:
        with private_creation_umask():
            browser_ctx = playwright.chromium.launch_persistent_context(
                user_data_dir=str(DOMO_PROFILE_DIR),
                headless=False,
                downloads_path=str(domo.TEMP_DOWNLOAD_DIR),
                accept_downloads=True,
            )
        page = browser_ctx.new_page()
        try:
            try:
                domo._authenticate(page, logger)
            except domo.PlaywrightTimeoutError:
                logger.error(
                    "  ✗ The private Domo session requires reauthentication. "
                    "The workflow did not pause. Run python3 auth_manager.py "
                    "--enroll-domo-keychain, then python3 auth_manager.py "
                    "--setup domo outside the release "
                    "run, then rerun with --reuse-domo-seeds."
                )
                return False
            for card in cards:
                output = card["output_fn"](ctx)
                logger.info(f"  Exporting {card['description']} → {output}")
                try:
                    domo._export_card(page, card, output, ctx, logger)
                    xlsx_output_fn = card.get("xlsx_output_fn")
                    if xlsx_output_fn:
                        xlsx_output = xlsx_output_fn(ctx)
                        convert_soundmouse_csv_to_xlsx(output, xlsx_output)
                        output.unlink()
                        logger.info(
                            f"  Converted CSV → upload-compatible XLSX: "
                            f"{xlsx_output.name}"
                        )
                except Exception as exc:  # browser errors are logged per card
                    logger.error(f"  ✗ {card['description']} failed: {exc}")
                    ok = False
        finally:
            browser_ctx.close()
            secure_private_directory(DOMO_PROFILE_DIR, recursive=True)
    return ok


def install_soundmouse_metadata(
    source_workbooks: list[Path],
    metadata_directory: Path,
    logger: logging.Logger,
) -> list[Path]:
    """Atomically install cleaned full-period workbooks in the delivery."""
    metadata_directory.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []
    for source in source_workbooks:
        destination = metadata_directory / source.name
        temporary = destination.with_name(f".{destination.name}.tmp")
        try:
            shutil.copy2(source, temporary)
            temporary.replace(destination)
        finally:
            if temporary.exists():
                temporary.unlink()
        outputs.append(destination)
        logger.info(f"  Installed metadata: {destination.name}")
    return outputs


def _wav_name(value: str) -> str:
    """Return a case-insensitive WAV leaf name for tracklist comparisons."""
    name = Path(str(value).strip()).name
    if name and not name.casefold().endswith(".wav"):
        name += ".wav"
    return name.casefold()


def _partition_soundmouse_rows(
    soundmouse_csv: Path,
    us_tracklist_csv: Path,
) -> tuple[list[str], list[dict[str, str]], list[dict[str, str]]]:
    """Split SoundMouse rows into US and non-US UniSync requests."""
    fields, rows = _read_csv(soundmouse_csv)
    filename_col = _find_column(fields, ["FILENAME", "FILE"])
    us_fields, us_rows = _read_csv(us_tracklist_csv)
    us_filename_col = _find_column(us_fields, ["FILENAME", "FILE"])
    if not filename_col or not us_filename_col:
        raise ValueError("SoundMouse and US tracklists both need a Filename column")

    us_names = {
        _wav_name(row.get(us_filename_col, ""))
        for row in us_rows
        if str(row.get(us_filename_col, "")).strip()
    }
    routed_us: list[dict[str, str]] = []
    routed_exus: list[dict[str, str]] = []
    for row in rows:
        destination = routed_us if _wav_name(row.get(filename_col, "")) in us_names else routed_exus
        destination.append(row)
    return fields, routed_us, routed_exus


def _write_soundmouse_request_csv(
    path: Path,
    fields: list[str],
    rows: list[dict[str, str]],
) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def _soundmouse_unisync_jobs(
    ctx: ReleaseContext,
    us_request_csv: Path | None = None,
    exus_request_csv: Path | None = None,
    all_request_csv: Path | None = None,
    destination_dir: Path | None = None,
) -> list[dict[str, str]]:
    """Build the additive territory jobs used by the SoundMouse delivery."""
    jobs = [
        {
            "name": f"SoundMouse {label} WAV ({ctx.soundmouse_activation_range})",
            "territory": territory,
            "cache_path": str(UPM_CACHE_WAV),
            "client_path": str(destination_dir or (ctx.soundmouse_release_dir / "MEDIA")),
            "csv": str(request_csv or all_request_csv or ctx.soundmouse_tracklist_csv),
        }
        for label, territory, request_csv in (
            ("US", "United States", us_request_csv),
            ("Ex-US", "Rest of World", exus_request_csv),
            ("Japan", "Japan", None),
        )
    ]
    jobs[1]["fallback_territory"] = "Japan"
    return jobs


def run_soundmouse_unisync(
    ctx: ReleaseContext,
    dry_run: bool,
    overwrite: bool,
    logger: logging.Logger,
    *,
    destination_dir: Path | None = None,
    existing_media_roots: tuple[Path, ...] = (),
) -> bool:
    """Run all three WAV territories into the workflow-period directory.

    SoundMouse's selected metadata workbooks can reference US, Rest-of-World,
    and Japan-only tracks.  A United States-only pass silently tops out at the
    US delivery count, so all territories must contribute to the same MEDIA
    folder.  UniSync's skip-existing behavior makes the later passes additive.
    """
    if dry_run:
        logger.info(
            "  [DRY RUN] Would route SoundMouse WAVs through US, "
            f"Rest of World, then Japan fallback → "
            f"{destination_dir or (ctx.soundmouse_release_dir / 'MEDIA')}"
        )
        return True

    from unisync_automation import STATUS_FAILED, run_all_unisync_jobs
    with tempfile.TemporaryDirectory(prefix="soundmouse_unisync_") as tmp:
        request_dir = Path(tmp)
        try:
            request_csv = ctx.soundmouse_tracklist_csv
            if existing_media_roots:
                fields, rows = _read_csv(request_csv)
                filename_col = _find_column(fields, POSSIBLE_FILENAME_COLS)
                if not filename_col:
                    raise ValueError("SoundMouse tracklist needs a Filename column")
                present = set().union(*(
                    _disk_file_keys(root) for root in existing_media_roots
                ))
                rows = [
                    row for row in rows
                    if _wav_name(row.get(filename_col, "")) not in present
                ]
                if not rows:
                    logger.info("  ✓ SoundMouse has no new audio to retrieve.")
                    return True
                request_csv = request_dir / "SoundMouse-Missing.csv"
                _write_soundmouse_request_csv(request_csv, fields, rows)
                logger.info(
                    f"  SoundMouse uploaded correction: {len(rows)} new audio row(s)."
                )
            fields, us_rows, exus_rows = _partition_soundmouse_rows(
                request_csv, ctx.us_tracklist_csv
            )
            us_csv = request_dir / "SoundMouse-US.csv"
            exus_csv = request_dir / "SoundMouse-ExUS.csv"
            _write_soundmouse_request_csv(us_csv, fields, us_rows)
            _write_soundmouse_request_csv(exus_csv, fields, exus_rows)
            logger.info(
                f"  SoundMouse territory routing: {len(us_rows)} US, "
                f"{len(exus_rows)} Rest-of-World/Japan fallback row(s)."
            )
            jobs = _soundmouse_unisync_jobs(
                ctx,
                us_csv,
                exus_csv,
                request_csv,
                destination_dir,
            )
        except (OSError, ValueError) as exc:
            logger.warning(
                f"  Could not partition SoundMouse requests ({exc}); "
                "falling back to full-list territory passes."
            )
            jobs = _soundmouse_unisync_jobs(
                ctx,
                all_request_csv=request_csv,
                destination_dir=destination_dir,
            )

        class _Jobs:
            unisync_jobs = jobs

        results = run_all_unisync_jobs(
            _Jobs(), dry_run=dry_run, logger=logger, overwrite=overwrite
        )
    return bool(results) and not any(v == STATUS_FAILED for v in results.values())


def download_soundmouse_covers(
    tracklist_csv: Path,
    release_directory: Path,
    dry_run: bool,
    overwrite: bool,
    logger: logging.Logger,
    *,
    existing_cover_roots: tuple[Path, ...] = (),
) -> bool:
    """Download unique covers into the workflow period's delivery root."""
    fields, rows = _read_csv(tracklist_csv)
    cover_col = _find_column(fields, POSSIBLE_COVER_COLS)
    url_col = _find_column(fields, POSSIBLE_URL_COLS)
    if not cover_col or not url_col:
        logger.error("  ✗ SoundMouse tracklist needs AlbumCoverArt and CDNAlbumArt columns.")
        return False

    if not dry_run:
        import requests

    ok = True
    seen: set[str] = set()
    existing_cover_names = set().union(*(
        _disk_file_keys(root) for root in existing_cover_roots
    )) if existing_cover_roots else set()
    for row in rows:
        name = Path(str(row.get(cover_col, "")).strip()).name
        url = str(row.get(url_col, "")).strip()
        if not name or name in seen:
            continue
        seen.add(name)
        destination = release_directory / "Covers" / name
        if name.casefold() in existing_cover_names:
            continue
        if destination.exists() and not overwrite:
            continue
        if not url:
            logger.warning(f"  No cover URL for {name}")
            ok = False
            continue
        if dry_run:
            logger.info(f"  [DRY RUN] Would download cover: {destination}")
            continue
        try:
            response = requests.get(url, timeout=45)
            response.raise_for_status()
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(response.content)
        except Exception as exc:
            logger.error(f"  ✗ Cover download failed ({name}): {exc}")
            ok = False
    logger.info(
        f"  SoundMouse covers for {release_directory.name}: {len(seen)}"
    )
    return ok


def _soundmouse_expected_names(tracklist_csv: Path) -> tuple[set[str], set[str]]:
    fields, rows = _read_csv(tracklist_csv)
    filename_col = _find_column(fields, POSSIBLE_FILENAME_COLS)
    cover_col = _find_column(fields, POSSIBLE_COVER_COLS)
    if not filename_col or not cover_col:
        raise ValueError(
            "SoundMouse tracklist needs Filename and AlbumCoverArt columns"
        )
    audio = {
        _wav_name(row.get(filename_col, ""))
        for row in rows
        if str(row.get(filename_col, "")).strip()
    }
    covers = {
        Path(str(row.get(cover_col, "")).strip()).name.casefold()
        for row in rows
        if str(row.get(cover_col, "")).strip()
    }
    return audio, covers


def _soundmouse_paths_by_leaf(root: Path) -> dict[str, list[Path]]:
    result: dict[str, list[Path]] = {}
    if not root.is_dir():
        return result
    for path in root.rglob("*"):
        if path.is_file() and path.name != ".DS_Store" and not path.name.startswith("._"):
            result.setdefault(path.name.casefold(), []).append(path)
    return result


def _archive_soundmouse_missing(path: Path, logger: logging.Logger) -> None:
    if not path.exists():
        return
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    archive = path.with_name(f"{path.name}-archived-{stamp}")
    counter = 2
    while archive.exists():
        archive = path.with_name(f"{path.name}-archived-{stamp}-{counter}")
        counter += 1
    path.replace(archive)
    logger.info(f"  Archived prior SoundMouse correction → {archive.name}")


def _write_soundmouse_correction_audit(
    path: Path,
    *,
    audio_additions: set[str],
    audio_removals: set[str],
    cover_additions: set[str],
    cover_removals: set[str],
) -> None:
    fields = ["Action", "Filename", "Local Result", "Required Manual Action"]
    rows: list[dict[str, str]] = []
    for action, values, local, manual in (
        (
            "AUDIO_ADDITION_UPLOAD", audio_additions,
            "Prepared under Missing/MEDIA", "Upload audio to SoundMouse",
        ),
        (
            "AUDIO_REMOVE_FROM_SOUNDMOUSE", audio_removals,
            "Removed from original MEDIA", "Remove audio from SoundMouse",
        ),
        (
            "COVER_ADDITION_UPLOAD", cover_additions,
            "Prepared under Missing/Covers", "Upload cover to SoundMouse",
        ),
        (
            "COVER_REMOVE_FROM_SOUNDMOUSE", cover_removals,
            "Removed from original Covers", "Remove cover from SoundMouse",
        ),
    ):
        rows.extend({
            "Action": action,
            "Filename": name,
            "Local Result": local,
            "Required Manual Action": manual,
        } for name in sorted(values))
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def run_soundmouse_step(
    ctx: ReleaseContext,
    dry_run: bool,
    overwrite: bool,
    logger: logging.Logger,
    *,
    reuse_domo_seeds: bool = False,
) -> bool:
    logger.info(f"  Tracklist: {ctx.soundmouse_tracklist_csv}")
    logger.info(f"  Release base: {SOUNDMOUSE_BASE}")
    from delivery_state import partner_needs_correction_package, partner_status

    status = partner_status(ctx.specials_dir, "soundmouse")
    correction_package = partner_needs_correction_package(
        ctx.specials_dir, "soundmouse"
    )
    logger.info(
        f"  SoundMouse refresh state: {status.upper()} — "
        + (
            "uploaded corrections use a separate Missing package."
            if correction_package
            else "refresh the original delivery in place."
        )
    )

    if dry_run:
        _export_domo_cards(ctx, _domo_configs(ctx), True, logger)
        create_soundmouse_directories_preview(ctx, logger)
        run_soundmouse_unisync(ctx, True, overwrite, logger)
        logger.info("  [DRY RUN] Covers will be derived from the exported tracklist.")
        logger.info("  [DRY RUN] Bucket-selected metadata (possible sheets):")
        _export_domo_cards(ctx, _domo_configs(ctx, list(METADATA_SHEETS)), True, logger)
        validate_soundmouse_delivery(
            [],
            ctx.soundmouse_release_dir / "MEDIA",
            ctx.soundmouse_release_dir / "Covers",
            ctx.soundmouse_validation_report,
            True,
            logger,
        )
        return True

    if reuse_domo_seeds:
        missing_seeds = [
            path for path in (
                ctx.soundmouse_tracklist_csv,
                ctx.soundmouse_bucket_csv,
            )
            if not path.is_file()
        ]
        if missing_seeds:
            logger.error(
                "  ✗ Cannot reuse SoundMouse Domo seed exports; missing: "
                + ", ".join(str(path) for path in missing_seeds)
            )
            return False
        logger.info("  ↩ Reusing existing SoundMouse tracklist and bucket exports.")
    elif not _export_domo_cards(ctx, _domo_configs(ctx), False, logger):
        return False
    try:
        root = create_soundmouse_directories(
            ctx.soundmouse_release_dir, False, logger
        )
    except (OSError, ValueError) as exc:
        logger.error(f"  ✗ SoundMouse directory setup failed: {exc}")
        return False
    try:
        expected_audio, expected_covers = _soundmouse_expected_names(
            ctx.soundmouse_tracklist_csv
        )
    except (OSError, ValueError) as exc:
        logger.error(f"  ✗ Could not read refreshed SoundMouse tracklist: {exc}")
        return False
    existing_audio = _soundmouse_paths_by_leaf(root / "MEDIA")
    existing_covers = _soundmouse_paths_by_leaf(root / "Covers")
    audio_additions = expected_audio - set(existing_audio)
    audio_removals = set(existing_audio) - expected_audio
    cover_additions = expected_covers - set(existing_covers)
    cover_removals = set(existing_covers) - expected_covers
    has_delta = bool(
        audio_additions or audio_removals or cover_additions or cover_removals
    )
    correction_root = root / "Missing"
    if correction_package and has_delta:
        _archive_soundmouse_missing(correction_root, logger)
        for child in (
            correction_root,
            correction_root / "MEDIA",
            correction_root / "Covers",
            correction_root / "Metadata",
        ):
            child.mkdir(parents=True, exist_ok=True)
    logger.info(
        f"  SoundMouse refresh delta: {len(audio_additions)} audio addition(s), "
        f"{len(audio_removals)} audio removal(s), "
        f"{len(cover_additions)} cover addition(s), "
        f"{len(cover_removals)} cover removal(s)."
    )
    try:
        codes = metadata_codes_from_bucket(ctx.soundmouse_bucket_csv)
    except (OSError, ValueError) as exc:
        logger.error(f"  ✗ Could not read SoundMouse bucket export: {exc}")
        return False
    if not codes:
        logger.error(
            "  ✗ SoundMouse bucket did not identify any known metadata sheet "
            "(01–10); refusing to guess."
        )
        return False
    logger.info(f"  SoundMouse metadata buckets: {', '.join(codes)}")
    with tempfile.TemporaryDirectory(prefix="soundmouse_metadata_") as tmp:
        staging_dir = Path(tmp)
        if not _export_domo_cards(
            ctx,
            _domo_configs(ctx, codes, metadata_dir=staging_dir),
            False,
            logger,
        ):
            return False
        source_workbooks = [
            staging_dir / metadata_filename(code) for code in codes
        ]
        media_destination = (
            correction_root / "MEDIA"
            if correction_package and has_delta
            else root / "MEDIA"
        )
        if not run_soundmouse_unisync(
            ctx,
            False,
            overwrite,
            logger,
            destination_dir=media_destination,
            existing_media_roots=(root / "MEDIA",) if correction_package else (),
        ):
            return False
        cover_destination = (
            correction_root
            if correction_package and has_delta
            else root
        )
        if not download_soundmouse_covers(
            ctx.soundmouse_tracklist_csv,
            cover_destination,
            False,
            overwrite,
            logger,
            existing_cover_roots=(root / "Covers",) if correction_package else (),
        ):
            return False
        try:
            metadata_paths = install_soundmouse_metadata(
                source_workbooks,
                root / "Metadata",
                logger,
            )
            if correction_package and has_delta:
                install_soundmouse_metadata(
                    source_workbooks,
                    correction_root / "Metadata",
                    logger,
                )
        except (OSError, ValueError) as exc:
            logger.error(f"  ✗ Could not install SoundMouse metadata: {exc}")
            return False

    prepared_audio = set().union(
        _disk_file_keys(root / "MEDIA"),
        _disk_file_keys(correction_root / "MEDIA")
        if correction_package and has_delta else set(),
    )
    prepared_covers = set().union(
        _disk_file_keys(root / "Covers"),
        _disk_file_keys(correction_root / "Covers")
        if correction_package and has_delta else set(),
    )
    unprepared_audio = expected_audio - prepared_audio
    unprepared_covers = expected_covers - prepared_covers
    if unprepared_audio or unprepared_covers:
        logger.error(
            "  ✗ SoundMouse refresh is incomplete; refusing to remove any "
            f"obsolete originals ({len(unprepared_audio)} audio and "
            f"{len(unprepared_covers)} cover additions still missing)."
        )
        return False

    # Only mutate the original delivery after all additions, covers, and
    # refreshed metadata have been prepared successfully.
    removal_errors = 0
    for key in audio_removals:
        for path in existing_audio[key]:
            try:
                path.unlink()
            except OSError as exc:
                logger.error(f"  ✗ Could not remove stale SoundMouse audio {path}: {exc}")
                removal_errors += 1
    for key in cover_removals:
        for path in existing_covers[key]:
            try:
                path.unlink()
            except OSError as exc:
                logger.error(f"  ✗ Could not remove stale SoundMouse cover {path}: {exc}")
                removal_errors += 1
    if removal_errors:
        return False
    if correction_package and has_delta:
        _write_soundmouse_correction_audit(
            correction_root / "SoundMouse Missing Audit.csv",
            audio_additions=audio_additions,
            audio_removals=audio_removals,
            cover_additions=cover_additions,
            cover_removals=cover_removals,
        )
        logger.warning(
            "  ⚠ SoundMouse was already uploaded. Review the Missing audit "
            "for manual removals in SoundMouse."
        )

    return validate_soundmouse_delivery(
        metadata_paths,
        root / "MEDIA",
        root / "Covers",
        ctx.soundmouse_validation_report,
        False,
        logger,
        additional_media_roots=(correction_root / "MEDIA",)
        if correction_package and has_delta else (),
        additional_cover_roots=(correction_root / "Covers",)
        if correction_package and has_delta else (),
    )


def create_soundmouse_directories_preview(
    ctx: ReleaseContext, logger: logging.Logger
) -> None:
    for child in (
        ctx.soundmouse_release_dir,
        ctx.soundmouse_release_dir / "Covers",
        ctx.soundmouse_release_dir / "Metadata",
        ctx.soundmouse_release_dir / "MEDIA",
    ):
        logger.info(f"  [DRY RUN] Would create: {child}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run SoundMouse Step 16")
    parser.add_argument("--year", type=int)
    parser.add_argument("--month", type=int)
    parser.add_argument("--part", type=int, choices=[1, 2])
    parser.add_argument("--previous-month", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument(
        "--reuse-domo-seeds",
        action="store_true",
        help=(
            "Reuse the existing SoundMouse tracklist and bucket CSVs; "
            "metadata cards are still refreshed."
        ),
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        ctx = context_from_cli_args(args)
    except ValueError as exc:
        print(f"ERROR: {exc}")
        return 1
    logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
    return 0 if run_soundmouse_step(
        ctx,
        args.dry_run,
        args.overwrite,
        logging.getLogger("soundmouse"),
        reuse_domo_seeds=args.reuse_domo_seeds,
    ) else 1


if __name__ == "__main__":
    raise SystemExit(main())
