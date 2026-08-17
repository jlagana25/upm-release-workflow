"""Offline synthetic-filesystem tests for SoundMouse Step 16."""

from __future__ import annotations

import csv
import logging
import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile

from config import ReleaseContext
from soundmouse import (
    _assert_soundmouse_xlsx_compatibility,
    _domo_configs,
    _partition_soundmouse_rows,
    _soundmouse_unisync_jobs,
    activation_ranges_from_tracklist,
    create_soundmouse_directories,
    convert_soundmouse_csv_to_xlsx,
    download_soundmouse_covers,
    install_soundmouse_correction_metadata,
    install_soundmouse_metadata,
    metadata_codes_from_bucket,
    validate_soundmouse_delivery,
)


class SoundMouseTests(unittest.TestCase):
    def _csv(self, root: Path, name: str, fields: list[str], rows: list[dict]) -> Path:
        path = root / name
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def test_context_names_full_month_and_parts(self) -> None:
        full = ReleaseContext(2026, 6, 1, previous_month=True)
        self.assertEqual(full.release_id, "UPM-2026-06-FULL")
        self.assertEqual(full.specials_root, "UPM-2026-06-FULL")
        self.assertEqual(full.hd_folder, "UPM-2026-06-FULL")
        self.assertEqual(full.month_display_folder, "June 2026 Full")
        self.assertEqual(
            full.us_tracklist_csv.name,
            "UPM-US-2026-06-FULL-Tracklist.csv",
        )
        self.assertEqual(
            full.pinned_cli_args(),
            ["--previous-month", "--year", "2026", "--month", "7"],
        )
        self.assertEqual(
            full.soundmouse_tracklist_csv.name,
            "Soundmouse 06-01-26 to 07-01-26.csv",
        )
        self.assertEqual(full.soundmouse_activation_range, "2026-06-01_to_2026-06-30")

        part1 = ReleaseContext(2026, 6, 1)
        part2 = ReleaseContext(2026, 6, 2)
        self.assertEqual(part1.release_id, "UPM-2026-06-P1")
        self.assertEqual(part2.release_id, "UPM-2026-06-P2")
        self.assertEqual(part1.month_display_folder, "June 2026 Part 1")
        self.assertEqual(part2.month_display_folder, "June 2026 Part 2")
        self.assertIn(
            "Universal Production Music June 2026 Part 1 - NBC",
            str(part1.partner_dirs["nbc_music_root"]),
        )
        self.assertEqual(
            part2.pinned_cli_args(),
            ["--year", "2026", "--month", "6", "--part", "2"],
        )
        self.assertEqual(part1.soundmouse_activation_range, "2026-06-01_to_2026-06-14")
        self.assertEqual(part2.soundmouse_activation_range, "2026-06-15_to_2026-06-30")
        self.assertEqual(part1.soundmouse_tracklist_csv.name, "Soundmouse 06-01-26 to 06-15-26.csv")
        self.assertEqual(part2.soundmouse_tracklist_csv.name, "Soundmouse 06-15-26 to 07-01-26.csv")

    def test_transition_full_month_and_rolling_delivery_names(self) -> None:
        transition_part1 = ReleaseContext(2026, 7, 1, previous_month=True)
        self.assertEqual(transition_part1.release_id, "UPM-2026-07-FULL")
        self.assertEqual(transition_part1.release_start, "2026-07-01")
        self.assertEqual(transition_part1.release_end, "2026-07-31")
        self.assertEqual(
            transition_part1.partner_folder_name("NBC"),
            "Universal Production Music August 2026 Part 1 - NBC",
        )
        self.assertEqual(
            transition_part1.partner_metadata["sourceaudio"].name,
            "UPM August 2026 Part 1 Metadata.csv",
        )

        transition = ReleaseContext(2026, 8, 2, full_month_content=True)
        self.assertEqual(transition.release_start, "2026-08-01")
        self.assertEqual(transition.release_end, "2026-08-31")
        self.assertEqual(
            transition.partner_folder_name("NBC"),
            "Universal Production Music August 2026 Part 2 - NBC",
        )
        self.assertIn("--full-month-content", transition.pinned_cli_args())

        same_month = ReleaseContext.for_date_range("2026-09-01", "2026-09-14")
        crossing = ReleaseContext.for_date_range("2026-09-29", "2026-10-12")
        self.assertEqual(
            same_month.partner_folder_name("NBC"),
            "Universal Production Music September 1–14 2026 Releases - NBC",
        )
        self.assertEqual(
            crossing.partner_folder_name("Japan NTT DATA"),
            "Universal Production Music September 29–October 12 2026 Releases - Japan NTT DATA",
        )
        self.assertEqual(
            crossing.pinned_cli_args(),
            ["--start-date", "2026-09-29", "--end-date", "2026-10-12"],
        )

        with self.assertRaisesRegex(ValueError, "exactly 14"):
            ReleaseContext.for_date_range("2026-09-01", "2026-09-15")

    def test_soundmouse_audio_uses_all_three_territories(self) -> None:
        ctx = ReleaseContext(2026, 6, 1, previous_month=True)
        jobs = _soundmouse_unisync_jobs(ctx)
        self.assertEqual(
            [job["territory"] for job in jobs],
            ["United States", "Rest of World", "Japan"],
        )
        self.assertEqual(len({job["client_path"] for job in jobs}), 1)
        self.assertTrue(jobs[0]["client_path"].endswith("2026-06-01_to_2026-06-30/MEDIA"))
        self.assertEqual(jobs[1]["fallback_territory"], "Japan")

    def test_soundmouse_rows_are_partitioned_by_us_tracklist(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            soundmouse = self._csv(
                root,
                "soundmouse.csv",
                ["Filename", "workAudioId"],
                [
                    {"Filename": "US_One", "workAudioId": "1"},
                    {"Filename": "EXUS_Two.wav", "workAudioId": "2"},
                    {"Filename": "US_Three.WAV", "workAudioId": "3"},
                ],
            )
            us = self._csv(
                root,
                "us.csv",
                ["Filename"],
                [{"Filename": "US_One.wav"}, {"Filename": "US_Three"}],
            )
            fields, us_rows, exus_rows = _partition_soundmouse_rows(soundmouse, us)
            self.assertEqual(fields, ["Filename", "workAudioId"])
            self.assertEqual([row["workAudioId"] for row in us_rows], ["1", "3"])
            self.assertEqual([row["workAudioId"] for row in exus_rows], ["2"])

    def test_directory_comes_from_workflow_period(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            release = root / "SoundMouse" / "2026-06-01_to_2026-06-30"
            created = create_soundmouse_directories(
                release, False, logging.getLogger("test")
            )
            self.assertEqual(created, release)
            for child in ("Covers", "Metadata", "MEDIA"):
                self.assertTrue((release / child).is_dir())

    def test_invalid_activation_range_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            tracklist = self._csv(
                root, "bad.csv", ["ActivationRange"],
                [{"ActivationRange": "../../escape"}],
            )
            with self.assertRaises(ValueError):
                activation_ranges_from_tracklist(tracklist)

    def test_bucket_recognizes_codes_and_territory_sets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            bucket = self._csv(
                root, "bucket.csv", ["SoundMouse Bucket", "Territory List"],
                [
                    {"SoundMouse Bucket": "02 UK DE SE OZ", "Territory List": ""},
                    {"SoundMouse Bucket": "", "Territory List": "US OZ"},
                    {"SoundMouse Bucket": "08 - OZ", "Territory List": ""},
                ],
            )
            self.assertEqual(metadata_codes_from_bucket(bucket), ["02", "07", "08"])

    def test_metadata_cards_export_csv_then_target_xlsx(self) -> None:
        ctx = ReleaseContext(2026, 7, 1, previous_month=True)
        with tempfile.TemporaryDirectory() as tmp:
            card = _domo_configs(ctx, ["01"], metadata_dir=Path(tmp))[0]
            self.assertEqual(card["format"], "csv")
            self.assertEqual(card["download_format"], "csv")
            self.assertEqual(card["output_fn"](ctx).suffix, ".csv")
            self.assertEqual(card["xlsx_output_fn"](ctx).suffix, ".xlsx")

    def test_csv_conversion_validates_shape_and_installs_xlsx(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_path = root / "metadata.csv"
            xlsx_path = root / "metadata.xlsx"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["Code", "Title", "Formula-like", "Long ID"])
                writer.writerow(["001", "Text, with comma", "=1+1", "12345678901234567890"])
                writer.writerow(["002", "", "literal", ""])

            convert_soundmouse_csv_to_xlsx(csv_path, xlsx_path)
            _assert_soundmouse_xlsx_compatibility(xlsx_path)

            with ZipFile(xlsx_path) as package:
                rewritten_parts = (
                    "[Content_Types].xml",
                    "xl/_rels/workbook.xml.rels",
                    "xl/sharedStrings.xml",
                    "xl/worksheets/sheet1.xml",
                )
                for name in rewritten_parts:
                    payload = package.read(name)
                    self.assertNotIn(b"ns0:", payload)
                    self.assertNotIn(b"xmlns:ns0=", payload)

            workbook = load_workbook(xlsx_path, data_only=False)
            sheet = workbook["Metadata"]
            self.assertEqual(sheet["A2"].value, "001")
            self.assertEqual(sheet["B2"].value, "Text, with comma")
            self.assertEqual(sheet["C2"].value, "=1+1")
            self.assertEqual(sheet["D2"].value, "12345678901234567890")
            self.assertEqual(sheet["A3"].value, "002")
            self.assertIsNone(sheet["B3"].value)
            self.assertEqual(sheet["C3"].value, "literal")
            self.assertIsNone(sheet["D3"].value)
            workbook.close()

    def test_csv_conversion_rejects_ragged_rows_before_excel(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_path = root / "bad.csv"
            csv_path.write_text("A,B\n1\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "row 2"):
                convert_soundmouse_csv_to_xlsx(csv_path, root / "bad.xlsx")

    def test_soundmouse_rejects_openpyxl_inline_string_packages(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "inline-strings.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "SoundMouse header"
            workbook.save(path)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "shared-string table"):
                _assert_soundmouse_xlsx_compatibility(path)

    def test_metadata_validation_checks_audio_and_covers(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            metadata_dir = root / "Metadata"
            media_dir = root / "MEDIA" / "Label" / "Album"
            covers_dir = root / "Covers"
            metadata_dir.mkdir()
            media_dir.mkdir(parents=True)
            covers_dir.mkdir()

            metadata = metadata_dir / "SoundMouseMetadata 01 - ALL.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Filename", "ALBUM ARTWORK FILE NAME"])
            sheet.append(["Track_One.wav", "cover-one.jpg"])
            sheet.append(["Track_Two.wav", "cover-two.jpg"])
            workbook.save(metadata)
            workbook.close()

            (media_dir / "TRACK_ONE.WAV").write_bytes(b"audio")
            (covers_dir / "Cover-One.JPG").write_bytes(b"cover")
            report = root / "reports" / "missing.csv"
            logger = logging.getLogger("test")

            self.assertFalse(validate_soundmouse_delivery(
                [metadata], root / "MEDIA", covers_dir, report, False, logger
            ))
            with report.open(encoding="utf-8-sig", newline="") as handle:
                missing = list(csv.DictReader(handle))
            self.assertEqual(
                {(row["Type"], row["Filename"]) for row in missing},
                {("AUDIO", "Track_Two.wav"), ("COVER", "cover-two.jpg")},
            )

            (media_dir / "Track_Two.wav").write_bytes(b"audio")
            (covers_dir / "cover-two.jpg").write_bytes(b"cover")
            self.assertTrue(validate_soundmouse_delivery(
                [metadata], root / "MEDIA", covers_dir, report, False, logger
            ))
            with report.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(list(csv.DictReader(handle)), [])

    def test_metadata_validation_accepts_uploaded_missing_package_union(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            metadata = root / "Metadata.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Filename", "ALBUM ARTWORK FILE NAME"])
            sheet.append(["old.wav", "old.jpg"])
            sheet.append(["new.wav", "new.jpg"])
            workbook.save(metadata)
            workbook.close()
            (root / "MEDIA").mkdir()
            (root / "Covers").mkdir()
            (root / "Missing" / "MEDIA").mkdir(parents=True)
            (root / "Missing" / "Covers").mkdir(parents=True)
            (root / "MEDIA" / "old.wav").write_bytes(b"old")
            (root / "Covers" / "old.jpg").write_bytes(b"old")
            (root / "Missing" / "MEDIA" / "new.wav").write_bytes(b"new")
            (root / "Missing" / "Covers" / "new.jpg").write_bytes(b"new")
            self.assertTrue(validate_soundmouse_delivery(
                [metadata],
                root / "MEDIA",
                root / "Covers",
                root / "report.csv",
                False,
                logging.getLogger("test"),
                additional_media_roots=(root / "Missing" / "MEDIA",),
                additional_cover_roots=(root / "Missing" / "Covers",),
            ))

    def test_uploaded_audio_correction_omits_unchanged_album_cover(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_csv = self._csv(
                root,
                "metadata.csv",
                ["Filename", "ALBUM ARTWORK FILE NAME", "Title"],
                [
                    {"Filename": "old.wav", "ALBUM ARTWORK FILE NAME": "shared.jpg", "Title": "Old"},
                    {"Filename": "new.wav", "ALBUM ARTWORK FILE NAME": "shared.jpg", "Title": "New"},
                ],
            )
            source_xlsx = root / "SoundMouseMetadata 01 - ALL.xlsx"
            convert_soundmouse_csv_to_xlsx(source_csv, source_xlsx)
            correction = root / "Missing"

            outputs = install_soundmouse_correction_metadata(
                [source_xlsx],
                correction / "Metadata",
                {"new.wav"},
                set(),
                logging.getLogger("test"),
            )
            self.assertEqual([path.name for path in outputs], [source_xlsx.name])
            _assert_soundmouse_xlsx_compatibility(outputs[0])
            workbook = load_workbook(outputs[0], read_only=True)
            values = list(workbook.active.iter_rows(values_only=True))
            workbook.close()
            self.assertEqual([row[0] for row in values], ["Filename", "new.wav"])

            tracklist = self._csv(
                root,
                "tracklist.csv",
                ["Filename", "AlbumCoverArt", "CDNAlbumArt"],
                [
                    {"Filename": "old.wav", "AlbumCoverArt": "shared.jpg", "CDNAlbumArt": ""},
                    {"Filename": "new.wav", "AlbumCoverArt": "shared.jpg", "CDNAlbumArt": ""},
                ],
            )
            original_covers = root / "Covers"
            original_covers.mkdir()
            (original_covers / "shared.jpg").write_bytes(b"cover")
            self.assertTrue(download_soundmouse_covers(
                tracklist,
                correction,
                False,
                False,
                logging.getLogger("test"),
                only_audio_names=None,
                only_cover_names=set(),
            ))
            self.assertFalse((correction / "Covers").exists())

    def test_full_month_metadata_is_installed_without_range_splitting(self) -> None:
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "SoundMouseMetadata 01 - ALL.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Filename", "ALBUM ARTWORK FILE NAME", "Formula"])
            sheet.append(["First.wav", "first.jpg", "=1+1"])
            sheet.append(["Second.wav", "second.jpg", "=2+2"])
            workbook.save(source)
            workbook.close()

            outputs = install_soundmouse_metadata(
                [source], root / "delivery" / "Metadata", logging.getLogger("test")
            )
            self.assertEqual(len(outputs), 1)
            installed = load_workbook(outputs[0], data_only=False)
            values = list(installed.active.iter_rows(values_only=True))
            installed.close()
            self.assertEqual(values[1], ("First.wav", "first.jpg", "=1+1"))
            self.assertEqual(values[2], ("Second.wav", "second.jpg", "=2+2"))


if __name__ == "__main__":
    unittest.main()
